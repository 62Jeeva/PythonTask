from openpyxl import load_workbook



class Utility:

    @staticmethod
    def row_count(excel_location):
        workbook=load_workbook(excel_location)
        sheet = workbook.active
        return sheet.max_row
  
    @staticmethod
    def read_excel_row(excel_location,row,col):
        workbook=load_workbook(excel_location)
        sheet = workbook.active
        return sheet.cell(row=row, column=col).value


    @staticmethod
    def write_excel_row(excel_location,row,col,data):
        workbook=load_workbook(excel_location)
        sheet = workbook.active
        sheet.cell(row=row, column=col).value = data
        workbook.save(excel_location)

